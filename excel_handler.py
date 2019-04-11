import datetime

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

import utils

def LoadTable(xlsx_path, read_only = True):
    excel = load_workbook(xlsx_path, read_only)
    table = excel.active
    return table

def GetUsers(table):
    users = set()
    for row in range(table.max_row):
        actual_row = row + 1
        if actual_row == 1:
            continue
        id = table.cell(actual_row, 2).value
        nid = int(id)
        users.add(nid)
    return users

def GetDates(table):
    dates = set()
    for row in range(table.max_row):
        actual_row = row + 1
        str = table.cell(actual_row, 1).value
        dt = datetime.datetime.strptime(str, "%Y-%m-%d")
        d = dt.date()
        dates.add(d)
    return dates

def SaveToWorkbook(users, date_list, ulist):
    wb = Workbook()
    font = Font(color = colors.RED)
    morning = datetime.time(9)
    evening = datetime.time(18)
    for id in users:
        if id not in ulist:
            # 始终未打卡
            ws = wb.create_sheet(user_info.name)
            continue
        user_info = ulist[id]
        ws = wb.create_sheet(user_info.name)
        counter = 'A'
        for date in date_list:
            ws["%c1" % counter] = date
            ws.column_dimensions["%c" % counter].width = 11
            if date not in user_info.goals:
                # 无打卡记录
                ws["%c2" % counter] = '无记录'
                ws["%c3" % counter] = '无记录'
                ws["%c2" % counter].font = font
                ws["%c3" % counter].font = font
            else:
                goal = user_info.goals[date]
                ws["%c2" % counter] = utils.TimeToStr(goal.earliest)
                ws["%c3" % counter] = utils.TimeToStr(goal.last)
                if goal.earliest > morning:
                    ws["%c2" % counter].font = font
                if goal.last < evening:
                    ws["%c3" % counter].font = font
            counter = chr(ord(counter) + 1)
    return wb

def SaveToExcel(wb, original_xlsx_path):
    sheet_count = len(wb.worksheets)
    if sheet_count > 1:
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    save_path = utils.GetOutputPathFromInputPath(original_xlsx_path)
    wb.save(save_path)
